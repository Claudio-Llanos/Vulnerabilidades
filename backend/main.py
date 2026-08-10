from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker
from sqlalchemy import text, event
from pydantic import BaseModel
from pydantic_settings import BaseSettings
from typing import Optional
from datetime import date, datetime
from email_utils import mail_evento
import pandas as pd, re, io

class Settings(BaseSettings):
    database_url: str = "postgresql://vulnuser:vuln2024secure@db:5432/vulnerabilidades"
    class Config:
        env_file = ".env"

settings = Settings()
async_url = settings.database_url.replace("postgresql://", "postgresql+asyncpg://")
engine = create_async_engine(async_url, pool_size=10, max_overflow=20)

@event.listens_for(engine.sync_engine, 'connect')
def set_tz(dbapi_conn, connection_record):
    cursor = dbapi_conn.cursor()
    cursor.execute("SET timezone='America/Santiago'")
    cursor.close()

Session = async_sessionmaker(engine, expire_on_commit=False)

app = FastAPI(title="Vulnerabilidades GoC", version="4.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

async def db():
    async with Session() as s:
        yield s

# ── Modelos ────────────────────────────────────────────────────────────────────

class VulnIn(BaseModel):
    tipo: Optional[str] = "Vulnerabilidad"
    detalle: str
    subgerencia: Optional[str] = None
    area: Optional[str] = None
    solicitante: Optional[str] = None
    responsable_om: Optional[str] = None
    responsable_ing: Optional[str] = None
    prioridad: Optional[int] = 1
    fecha_declaracion: Optional[date] = None
    fecha_compromiso: Optional[date] = None
    fecha_solucion: Optional[date] = None
    estado_om: Optional[str] = "PENDIENTE"
    estado_ing: Optional[str] = None
    condicion_ing: Optional[str] = None
    obs_om: Optional[str] = None
    obs_ing: Optional[str] = None
    usuario_nombre: Optional[str] = None

class VulnPatch(BaseModel):
    tipo: Optional[str] = None
    detalle: Optional[str] = None
    subgerencia: Optional[str] = None
    area: Optional[str] = None
    solicitante: Optional[str] = None
    responsable_om: Optional[str] = None
    responsable_ing: Optional[str] = None
    prioridad: Optional[int] = None
    fecha_declaracion: Optional[date] = None
    fecha_compromiso: Optional[date] = None
    fecha_solucion: Optional[date] = None
    estado_om: Optional[str] = None
    estado_ing: Optional[str] = None
    condicion_ing: Optional[str] = None
    obs_om: Optional[str] = None
    obs_ing: Optional[str] = None
    usuario_nombre: Optional[str] = None
    requiere_capex: Optional[bool] = None
    estado_capex: Optional[str] = None

class NotaIn(BaseModel):
    autor: Optional[str] = "Usuario"
    texto: str

class CfgItem(BaseModel):
    nombre: str

class CfgArea(BaseModel):
    nombre: str
    subgerencia: Optional[str] = None

class LoginIn(BaseModel):
    username: str
    password: str

class UsuarioIn(BaseModel):
    username: str
    nombre: str
    password: str
    rol: str = "editor"

class UsuarioPatch(BaseModel):
    nombre: Optional[str] = None
    password: Optional[str] = None
    rol: Optional[str] = None
    activo: Optional[bool] = None

class MailDestIn(BaseModel):
    nombre: Optional[str] = None
    email: str

# ── Helpers ────────────────────────────────────────────────────────────────────

def capitalize_text(val):
    if not val: return val
    s = str(val).strip()
    return s[0].upper() + s[1:] if s else s

def safe(val):
    if val is None: return None
    if isinstance(val, float) and pd.isna(val): return None
    s = str(val).strip()
    return s if s and s.lower() not in ('nan','none','nat') else None

def title_case(val):
    v = safe(val)
    return v.title() if v else None

def safe_date(val):
    if val is None: return None
    if isinstance(val, float) and pd.isna(val): return None
    s = str(val).strip().upper()
    if any(x in s for x in ['SIN','FECHA','COMPROMISO','N/A']): return None
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) else val.date()
    try: return pd.to_datetime(val).date()
    except: return None

def safe_int(val):
    try: return int(float(val))
    except: return None

def norm_estado(val):
    if not val: return "SIN ACCION"
    v = str(val).strip().upper()
    return v if v in ['SIN ACCION','EN CURSO','FINALIZADO'] else "SIN ACCION"

def norm_subgerencia(val):
    if not val: return None
    v = str(val).strip().replace('\n',' ')
    return v.upper() if v and v.lower() not in ('nan','none') else None

def split_detalle_obs(raw):
    if not raw: return None, None
    texto = str(raw).strip()
    patterns = [r'\n[Rr]esp[a-z]*\.?\s+[Oo][yY&]\s*[Mm]\s*[:.]\s*', r'\nRespuesta\s*[:.]\s*']
    for p in patterns:
        m = re.search(p, texto)
        if m:
            return texto[:m.start()].strip() or texto, texto[m.end():].strip() or None
    return texto, None

async def get_mail_dest(s: AsyncSession):
    rows = await s.execute(text("SELECT email FROM mail_destinatarios WHERE activo=TRUE"))
    return [r[0] for r in rows]

# ── AUTH ───────────────────────────────────────────────────────────────────────

@app.post("/api/login")
async def login(body: LoginIn, s: AsyncSession = Depends(db)):
    row = await s.execute(
        text("SELECT id, username, nombre, rol, activo FROM usuarios WHERE username = :u AND password = :p"),
        {"u": body.username, "p": body.password}
    )
    user = row.fetchone()
    if not user:
        raise HTTPException(401, "Usuario o contraseña incorrectos")
    if not user.activo:
        raise HTTPException(403, "Usuario desactivado")
    return {"id": user.id, "username": user.username, "nombre": user.nombre, "rol": user.rol}

# ── HEALTH ─────────────────────────────────────────────────────────────────────

@app.get("/health")
async def health(): return {"status": "ok", "version": "4.0"}

# ── DASHBOARD ──────────────────────────────────────────────────────────────────

@app.get("/api/dashboard")
async def dashboard(s: AsyncSession = Depends(db)):
    total       = await s.execute(text("SELECT COUNT(*) FROM vulnerabilidades"))
    estados     = await s.execute(text("SELECT estado_om, COUNT(*) FROM vulnerabilidades GROUP BY estado_om ORDER BY COUNT(*) DESC"))
    prioridades = await s.execute(text("SELECT prioridad, COUNT(*) FROM vulnerabilidades WHERE prioridad IS NOT NULL GROUP BY prioridad ORDER BY prioridad"))
    subgs       = await s.execute(text("SELECT subgerencia, COUNT(*) FROM vulnerabilidades WHERE subgerencia IS NOT NULL GROUP BY subgerencia ORDER BY COUNT(*) DESC LIMIT 10"))
    responsables= await s.execute(text("SELECT responsable_ing, COUNT(*) FROM vulnerabilidades WHERE responsable_ing IS NOT NULL GROUP BY responsable_ing ORDER BY COUNT(*) DESC LIMIT 10"))
    vencidas    = await s.execute(text("SELECT COUNT(*) FROM vulnerabilidades WHERE fecha_compromiso < NOW() AND estado_om NOT IN ('FINALIZADO')"))
    crit_data   = await s.execute(text("SELECT prioridad, estado_om, COUNT(*) FROM vulnerabilidades WHERE prioridad IS NOT NULL GROUP BY prioridad, estado_om ORDER BY prioridad, estado_om"))
    return {
        "total": total.scalar(), "vencidas": vencidas.scalar(),
        "por_estado": [{"estado":r[0],"total":r[1]} for r in estados],
        "por_prioridad": [{"prioridad":r[0],"total":r[1]} for r in prioridades],
        "por_subgerencia": [{"subgerencia":r[0],"total":r[1]} for r in subgs],
        "por_responsable": [{"responsable":r[0],"total":r[1]} for r in responsables],
        "criticidad_estado": [{"prioridad":r[0],"estado":r[1],"total":r[2]} for r in crit_data],
    }

# ── VULNERABILIDADES ───────────────────────────────────────────────────────────

@app.get("/api/vulnerabilidades")
async def listar(
    s: AsyncSession = Depends(db),
    page: int = Query(1,ge=1), limit: int = Query(50,ge=1,le=2000),
    q: Optional[str]=None, estado: Optional[str]=None,
    prioridad: Optional[int]=None, subgerencia: Optional[str]=None,
    area: Optional[str]=None, tipo: Optional[str]=None,
    responsable_ing: Optional[str]=None, sort: str="id", dir: str="asc",
):
    where, params = [], {}
    if q:
        where.append("(detalle ILIKE :q OR area ILIKE :q OR responsable_ing ILIKE :q OR subgerencia ILIKE :q OR tipo ILIKE :q)")
        params["q"] = f"%{q}%"
    if estado:   where.append("estado_om = :estado");         params["estado"] = estado
    if prioridad:where.append("prioridad = :prioridad");      params["prioridad"] = prioridad
    if subgerencia:where.append("subgerencia = :subgerencia");params["subgerencia"] = subgerencia
    if area:     where.append("area ILIKE :area");            params["area"] = f"%{area}%"
    if tipo:     where.append("tipo ILIKE :tipo");            params["tipo"] = f"%{tipo}%"
    if responsable_ing: where.append("responsable_ing ILIKE :resp"); params["resp"] = f"%{responsable_ing}%"
    w  = ("WHERE " + " AND ".join(where)) if where else ""
    safe_cols = {"id","prioridad","estado_om","subgerencia","area","tipo","fecha_declaracion","fecha_solucion","updated_at"}
    sc = sort if sort in safe_cols else "id"
    sd = "DESC" if dir.lower()=="desc" else "ASC"
    offset = (page-1)*limit
    cnt  = await s.execute(text(f"SELECT COUNT(*) FROM vulnerabilidades {w}"), params)
    rows = await s.execute(text(f"""
        SELECT id, tipo, detalle, subgerencia, area, solicitante,
               responsable_om, responsable_ing, prioridad,
               fecha_declaracion, fecha_compromiso, fecha_solucion,
               estado_om, estado_ing, condicion_ing, updated_at
        FROM vulnerabilidades {w}
        ORDER BY {sc} {sd} LIMIT :limit OFFSET :offset
    """), {**params, "limit":limit, "offset":offset})
    return {"total":cnt.scalar(),"page":page,"limit":limit,"items":[dict(r._mapping) for r in rows]}

@app.get("/api/vulnerabilidades/{vid}")
async def obtener(vid: int, s: AsyncSession = Depends(db)):
    row = await s.execute(text("SELECT * FROM vulnerabilidades WHERE id = :id"), {"id":vid})
    r = row.fetchone()
    if not r: raise HTTPException(404,"No encontrado")
    vuln = dict(r._mapping)
    notas = await s.execute(text("SELECT * FROM notas WHERE vulnerabilidad_id = :id ORDER BY created_at DESC"), {"id":vid})
    hist  = await s.execute(text("SELECT * FROM historial WHERE vulnerabilidad_id = :id ORDER BY created_at DESC LIMIT 30"), {"id":vid})
    vuln["notas"]     = [dict(n._mapping) for n in notas]
    vuln["historial"] = [dict(h._mapping) for h in hist]
    return vuln

@app.post("/api/vulnerabilidades", status_code=201)
async def crear(body: VulnIn, s: AsyncSession = Depends(db)):
    usuario = body.usuario_nombre or "Usuario"
    d = {k:v for k,v in body.model_dump(exclude_none=True).items() if k not in ('usuario_nombre',)}
    for field in ['detalle','obs_om','obs_ing','condicion_ing']:
        if field in d and d[field]: d[field] = capitalize_text(d[field])
    cols = ", ".join(d.keys()); vals = ", ".join(f":{k}" for k in d.keys())
    res  = await s.execute(text(f"INSERT INTO vulnerabilidades ({cols}) VALUES ({vals}) RETURNING id"), d)
    new_id = res.scalar()
    await s.execute(text("INSERT INTO historial (vulnerabilidad_id,campo,valor_anterior,valor_nuevo,usuario) VALUES (:id,:c,:a,:n,:u)"),
        {"id":new_id,"c":"CREACION","a":"—","n":"Registro creado","u":usuario})
    await s.commit()
    dest = await get_mail_dest(s)
    import asyncio
    asyncio.create_task(mail_evento(
        evento="Nueva vulnerabilidad creada", vuln_id=new_id,
        detalle=d.get("detalle","")[:100], usuario=usuario,
        cambios=[f"Tipo: {d.get('tipo','')}",f"Area: {d.get('area','')}",
                 f"Prioridad: P{d.get('prioridad','')}",f"Estado: {d.get('estado_om','PENDIENTE')}"],
        dest=dest
    ))
    return {"id": new_id}

@app.patch("/api/vulnerabilidades/{vid}")
async def actualizar(vid: int, body: VulnPatch, s: AsyncSession = Depends(db)):
    row = await s.execute(text("SELECT * FROM vulnerabilidades WHERE id = :id"), {"id":vid})
    cur = row.fetchone()
    if not cur: raise HTTPException(404,"No encontrado")
    usuario = body.usuario_nombre or "Usuario"
    updates = {k:v for k,v in body.model_dump(exclude_none=True).items() if k not in ('usuario_nombre',)}
    for field in ['detalle','obs_om','obs_ing','condicion_ing']:
        if field in updates and updates[field]: updates[field] = capitalize_text(updates[field])
    if updates.get("estado_om") == "FINALIZADO" and not updates.get("fecha_solucion"):
        updates["fecha_solucion"] = date.today()
    if not updates: return {"message":"Sin cambios"}
    cur_dict = dict(cur._mapping)
    cambios_log = []
    for campo, nuevo in updates.items():
        anterior = cur_dict.get(campo)
        if str(anterior) != str(nuevo):
            await s.execute(text("INSERT INTO historial (vulnerabilidad_id,campo,valor_anterior,valor_nuevo,usuario) VALUES (:id,:c,:a,:n,:u)"),
                {"id":vid,"c":campo,"a":str(anterior),"n":str(nuevo),"u":usuario})
            cambios_log.append({"campo":campo,"anterior":str(anterior),"nuevo":str(nuevo)})
    set_sql = ", ".join(f"{k} = :{k}" for k in updates)
    await s.execute(text(f"UPDATE vulnerabilidades SET {set_sql} WHERE id = :id"), {**updates,"id":vid})
    await s.commit()
    if cambios_log:
        dest = await get_mail_dest(s)
        detalle_txt = cur_dict.get("detalle","")[:100]
        import asyncio
        asyncio.create_task(mail_evento(
            evento="Vulnerabilidad modificada", vuln_id=vid,
            detalle=detalle_txt, usuario=usuario,
            cambios=[f"{h['campo']}: {h['anterior']} -> {h['nuevo']}" for h in cambios_log],
            dest=dest
        ))
    return {"message":"OK"}

@app.delete("/api/vulnerabilidades/{vid}")
async def eliminar(vid: int, s: AsyncSession = Depends(db)):
    res = await s.execute(text("DELETE FROM vulnerabilidades WHERE id = :id RETURNING id"), {"id":vid})
    if not res.fetchone(): raise HTTPException(404,"No encontrado")
    await s.commit()
    return {"message":"Eliminado"}

@app.post("/api/vulnerabilidades/{vid}/notas", status_code=201)
async def agregar_nota(vid: int, body: NotaIn, s: AsyncSession = Depends(db)):
    row = await s.execute(text("SELECT detalle FROM vulnerabilidades WHERE id = :id"), {"id":vid})
    r = row.fetchone()
    if not r: raise HTTPException(404,"No encontrado")
    texto_norm = capitalize_text(body.texto) if body.texto else body.texto
    await s.execute(text("INSERT INTO notas (vulnerabilidad_id,autor,texto) VALUES (:id,:autor,:texto)"),
        {"id":vid,"autor":body.autor,"texto":texto_norm})
    await s.execute(text("INSERT INTO historial (vulnerabilidad_id,campo,valor_anterior,valor_nuevo,usuario) VALUES (:id,:c,:a,:n,:u)"),
        {"id":vid,"c":"NOTA","a":"—","n":texto_norm[:100],"u":body.autor})
    await s.commit()
    dest = await get_mail_dest(s)
    import asyncio
    asyncio.create_task(mail_evento(
        evento="Nota agregada", vuln_id=vid,
        detalle=r[0][:100], usuario=body.autor,
        cambios=[f"Nota: {texto_norm[:200]}"], dest=dest
    ))
    return {"message":"Nota agregada"}

# ── FILTROS Y CONFIG ────────────────────────────────────────────────────────────

@app.get("/api/filtros")
async def filtros(s: AsyncSession = Depends(db)):
    subgs  = await s.execute(text("SELECT nombre FROM cfg_subgerencias WHERE activo=TRUE ORDER BY nombre"))
    areas  = await s.execute(text("SELECT nombre, subgerencia FROM cfg_areas WHERE activo=TRUE ORDER BY subgerencia, nombre"))
    tipos  = await s.execute(text("SELECT nombre FROM cfg_tipos WHERE activo=TRUE ORDER BY nombre"))
    resps  = await s.execute(text("SELECT nombre FROM cfg_responsables WHERE activo=TRUE ORDER BY nombre"))
    return {
        "subgerencias": [r[0] for r in subgs],
        "areas": [{"nombre":r[0],"subgerencia":r[1]} for r in areas],
        "tipos": [r[0] for r in tipos],
        "responsables_ing": [r[0] for r in resps],
    }

@app.get("/api/config/subgerencias")
async def get_subgs(s: AsyncSession = Depends(db)):
    r = await s.execute(text("SELECT id,nombre,activo FROM cfg_subgerencias ORDER BY nombre"))
    return [dict(x._mapping) for x in r]

@app.post("/api/config/subgerencias", status_code=201)
async def add_subg(body: CfgItem, s: AsyncSession = Depends(db)):
    await s.execute(text("INSERT INTO cfg_subgerencias (nombre) VALUES (:n) ON CONFLICT DO NOTHING"), {"n":body.nombre.upper()})
    await s.commit(); return {"message":"OK"}

@app.delete("/api/config/subgerencias/{id}")
async def del_subg(id: int, s: AsyncSession = Depends(db)):
    await s.execute(text("UPDATE cfg_subgerencias SET activo=FALSE WHERE id=:id"), {"id":id})
    await s.commit(); return {"message":"OK"}

@app.get("/api/config/areas")
async def get_areas(s: AsyncSession = Depends(db)):
    r = await s.execute(text("SELECT id,nombre,subgerencia,activo FROM cfg_areas ORDER BY subgerencia,nombre"))
    return [dict(x._mapping) for x in r]

@app.post("/api/config/areas", status_code=201)
async def add_area(body: CfgArea, s: AsyncSession = Depends(db)):
    await s.execute(text("INSERT INTO cfg_areas (nombre,subgerencia) VALUES (:n,:s) ON CONFLICT DO NOTHING"),
        {"n":body.nombre,"s":body.subgerencia})
    await s.commit(); return {"message":"OK"}

@app.delete("/api/config/areas/{id}")
async def del_area(id: int, s: AsyncSession = Depends(db)):
    await s.execute(text("UPDATE cfg_areas SET activo=FALSE WHERE id=:id"), {"id":id})
    await s.commit(); return {"message":"OK"}

@app.get("/api/config/tipos")
async def get_tipos(s: AsyncSession = Depends(db)):
    r = await s.execute(text("SELECT id,nombre,activo FROM cfg_tipos ORDER BY nombre"))
    return [dict(x._mapping) for x in r]

@app.post("/api/config/tipos", status_code=201)
async def add_tipo(body: CfgItem, s: AsyncSession = Depends(db)):
    await s.execute(text("INSERT INTO cfg_tipos (nombre) VALUES (:n) ON CONFLICT DO NOTHING"), {"n":body.nombre})
    await s.commit(); return {"message":"OK"}

@app.delete("/api/config/tipos/{id}")
async def del_tipo(id: int, s: AsyncSession = Depends(db)):
    await s.execute(text("UPDATE cfg_tipos SET activo=FALSE WHERE id=:id"), {"id":id})
    await s.commit(); return {"message":"OK"}

@app.get("/api/config/responsables")
async def get_resps(s: AsyncSession = Depends(db)):
    r = await s.execute(text("SELECT id,nombre,activo FROM cfg_responsables ORDER BY nombre"))
    return [dict(x._mapping) for x in r]

@app.post("/api/config/responsables", status_code=201)
async def add_resp(body: CfgItem, s: AsyncSession = Depends(db)):
    await s.execute(text("INSERT INTO cfg_responsables (nombre) VALUES (:n) ON CONFLICT DO NOTHING"), {"n":body.nombre})
    await s.commit(); return {"message":"OK"}

@app.delete("/api/config/responsables/{id}")
async def del_resp(id: int, s: AsyncSession = Depends(db)):
    await s.execute(text("UPDATE cfg_responsables SET activo=FALSE WHERE id=:id"), {"id":id})
    await s.commit(); return {"message":"OK"}

# ── ADMIN: USUARIOS ────────────────────────────────────────────────────────────

@app.get("/api/admin/usuarios")
async def get_usuarios(s: AsyncSession = Depends(db)):
    r = await s.execute(text("SELECT id, username, nombre, rol, activo, created_at FROM usuarios ORDER BY rol, nombre"))
    return [dict(x._mapping) for x in r]

@app.post("/api/admin/usuarios", status_code=201)
async def crear_usuario(body: UsuarioIn, s: AsyncSession = Depends(db)):
    try:
        await s.execute(text(
            "INSERT INTO usuarios (username, nombre, password, rol) VALUES (:u, :n, :p, :r)"
        ), {"u": body.username, "n": body.nombre, "p": body.password, "r": body.rol})
        await s.commit()
        return {"message": "Usuario creado"}
    except Exception:
        raise HTTPException(400, "El nombre de usuario ya existe")

@app.patch("/api/admin/usuarios/{uid}")
async def actualizar_usuario(uid: int, body: UsuarioPatch, s: AsyncSession = Depends(db)):
    updates = {k:v for k,v in body.model_dump(exclude_none=True).items()}
    if not updates: return {"message": "Sin cambios"}
    set_sql = ", ".join(f"{k} = :{k}" for k in updates)
    await s.execute(text(f"UPDATE usuarios SET {set_sql} WHERE id = :id"), {**updates, "id": uid})
    await s.commit()
    return {"message": "Actualizado"}

@app.delete("/api/admin/usuarios/{uid}")
async def desactivar_usuario(uid: int, s: AsyncSession = Depends(db)):
    await s.execute(text("UPDATE usuarios SET activo=FALSE WHERE id=:id"), {"id": uid})
    await s.commit()
    return {"message": "Desactivado"}

# ── ADMIN: DESTINATARIOS CORREO ─────────────────────────────────────────────────

@app.get("/api/admin/mail")
async def get_mail(s: AsyncSession = Depends(db)):
    r = await s.execute(text("SELECT id, nombre, email, activo FROM mail_destinatarios ORDER BY nombre"))
    return [dict(x._mapping) for x in r]

@app.post("/api/admin/mail", status_code=201)
async def add_mail(body: MailDestIn, s: AsyncSession = Depends(db)):
    try:
        await s.execute(text(
            "INSERT INTO mail_destinatarios (nombre, email) VALUES (:n, :e)"
        ), {"n": body.nombre, "e": body.email})
        await s.commit()
        return {"message": "Destinatario agregado"}
    except Exception:
        raise HTTPException(400, "El correo ya existe")

@app.delete("/api/admin/mail/{mid}")
async def del_mail(mid: int, s: AsyncSession = Depends(db)):
    await s.execute(text("UPDATE mail_destinatarios SET activo=FALSE WHERE id=:id"), {"id": mid})
    await s.commit()
    return {"message": "Eliminado"}

# ── IMPORTAR ───────────────────────────────────────────────────────────────────

@app.post("/api/importar")
async def importar(file: UploadFile = File(...), s: AsyncSession = Depends(db)):
    raw = await file.read()
    try:
        eng = "xlrd" if file.filename.endswith(".xls") else "openpyxl"
        xl  = pd.ExcelFile(io.BytesIO(raw), engine=eng)
        inserted, skipped = 0, 0
        for sheet in xl.sheet_names:
            df = xl.parse(sheet, header=0)
            df.columns = [str(c).strip() for c in df.columns]
            df = df[[c for c in df.columns if not c.startswith("Unnamed")]]
            if "DETALLE" not in df.columns: continue
            for _, row in df.iterrows():
                try:
                    raw_det = safe(row.get("DETALLE")) or safe(row.get("TIPO"))
                    if not raw_det: skipped += 1; continue
                    detalle, obs_extra = split_detalle_obs(raw_det)
                    obs_col   = safe(row.get("OBSERVACIONES O&M"))
                    obs_final = "\n\n".join(filter(None, [obs_col, obs_extra])) or None
                    resp_ing  = title_case(row.get("RESPONSABLE ING"))
                    if resp_ing:
                        await s.execute(text("INSERT INTO cfg_responsables (nombre) VALUES (:n) ON CONFLICT DO NOTHING"), {"n": resp_ing})
                    params = {
                        "tipo":              title_case(row.get("TIPO")) or "Vulnerabilidad",
                        "detalle":           detalle,
                        "subgerencia":       norm_subgerencia(row.get("SUBGERENCIA")),
                        "area":              title_case(row.get("AREA")),
                        "solicitante":       title_case(row.get("SOLICITANTE")),
                        "responsable_om":    title_case(row.get("RESPONSABLE")),
                        "responsable_ing":   resp_ing,
                        "prioridad":         safe_int(row.get("PRIORIDAD")),
                        "fecha_declaracion": safe_date(row.get("FECHA DECLARACION")),
                        "fecha_compromiso":  safe_date(row.get("FECHA COMPROMISO")),
                        "fecha_solucion":    safe_date(row.get("FECHA SOLUCION")),
                        "estado_om":         norm_estado(row.get("ESTADO O&M")),
                        "estado_ing":        safe(row.get("ESTADO ING")),
                        "condicion_ing":     safe(row.get("CONDICION ING")),
                        "obs_om":            obs_final,
                        "obs_ing":           safe(row.get("OBSERVACIONES ING")),
                    }
                    sql = ("INSERT INTO vulnerabilidades "
                        "(tipo,detalle,subgerencia,area,solicitante,responsable_om,responsable_ing,"
                        "prioridad,fecha_declaracion,fecha_compromiso,fecha_solucion,"
                        "estado_om,estado_ing,condicion_ing,obs_om,obs_ing) VALUES "
                        "(:tipo,:detalle,:subgerencia,:area,:solicitante,:responsable_om,:responsable_ing,"
                        ":prioridad,:fecha_declaracion,:fecha_compromiso,:fecha_solucion,"
                        ":estado_om,:estado_ing,:condicion_ing,:obs_om,:obs_ing)")
                    await s.execute(text("SAVEPOINT sp_row"))
                    await s.execute(text(sql), params)
                    inserted += 1
                except Exception:
                    await s.execute(text("ROLLBACK TO SAVEPOINT sp_row"))
                    skipped += 1
        await s.commit()
        return {"inserted": inserted, "skipped": skipped}
    except Exception as e:
        raise HTTPException(400, f"Error: {e}")

from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

@app.get("/api/exportar")
async def exportar(
    s: AsyncSession = Depends(db),
    q: Optional[str] = None,
    estado: Optional[str] = None,
    prioridad: Optional[int] = None,
    subgerencia: Optional[str] = None,
    area: Optional[str] = None,
    tipo: Optional[str] = None,
):
    where, params = [], {}
    if q:
        where.append("(detalle ILIKE :q OR area ILIKE :q OR responsable_ing ILIKE :q OR subgerencia ILIKE :q OR tipo ILIKE :q)")
        params["q"] = f"%{q}%"
    if estado:      where.append("estado_om = :estado");          params["estado"] = estado
    if prioridad:   where.append("prioridad = :prioridad");        params["prioridad"] = prioridad
    if subgerencia: where.append("subgerencia = :subgerencia");    params["subgerencia"] = subgerencia
    if area:        where.append("area ILIKE :area");              params["area"] = f"%{area}%"
    if tipo:        where.append("tipo ILIKE :tipo");              params["tipo"] = f"%{tipo}%"
    w = ("WHERE " + " AND ".join(where)) if where else ""
    rows = await s.execute(text(f"""
        SELECT id, subgerencia, area, tipo, detalle, prioridad,
               estado_om, responsable_ing, fecha_declaracion,
               fecha_solucion, obs_om, obs_ing
        FROM vulnerabilidades {w}
        ORDER BY id ASC
    """), params)
    data = rows.fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vulnerabilidades"
    header_fill = PatternFill("solid", fgColor="DA0812")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    headers = ["ID","Subgerencia","Area","Tipo","Detalle","Prioridad","Estado","Resp. ING","Declaracion","Cierre","Obs. O&M","Obs. ING"]
    col_widths = [6,30,15,20,50,12,14,20,18,18,40,40]
    for col, (h, w2) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w2
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    pri_labels = {1:"P1 - Critica", 2:"P2 - Media", 3:"P3 - Baja"}
    for i, row in enumerate(data, 2):
        vals = [row[0], row[1], row[2], row[3], row[4],
                pri_labels.get(row[5], str(row[5] or "")),
                row[6], row[7],
                str(row[8]) if row[8] else "",
                str(row[9]) if row[9] else "",
                row[10] or "", row[11] or ""]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FFF5F5")
        ws.row_dimensions[i].height = 40
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from datetime import date as dt
    fname = f"vulnerabilidades_{dt.today()}.xlsx"
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

def norm_columnas(df):
    """Normaliza nombres de columnas para soportar ambos formatos"""
    mapa = {
        'detalle': 'DETALLE',
        'subgerencia': 'SUBGERENCIA',
        'area': 'AREA',
        'tipo': 'TIPO',
        'prioridad': 'PRIORIDAD',
        'estado': 'ESTADO O&M',
        'estado o&m': 'ESTADO O&M',
        'resp. ing': 'RESPONSABLE ING',
        'responsable ing': 'RESPONSABLE ING',
        'declaracion': 'FECHA DECLARACION',
        'fecha declaracion': 'FECHA DECLARACION',
        'cierre': 'FECHA SOLUCION',
        'fecha solucion': 'FECHA SOLUCION',
        'obs. o&m': 'OBSERVACIONES O&M',
        'observaciones o&m': 'OBSERVACIONES O&M',
        'obs. ing': 'OBSERVACIONES ING',
        'observaciones ing': 'OBSERVACIONES ING',
    }
    df.columns = [mapa.get(str(c).strip().lower(), str(c).strip().upper()) for c in df.columns]
    return df

def norm_prioridad(val):
    if val is None: return None
    s = str(val).strip().upper()
    if 'P1' in s or s == '1': return 1
    if 'P2' in s or s == '2': return 2
    if 'P3' in s or s == '3': return 3
    try: return int(float(val))
    except: return None

@app.post("/api/importar2")
async def importar2(file: UploadFile = File(...), s: AsyncSession = Depends(db)):
    raw = await file.read()
    try:
        eng = "xlrd" if file.filename.endswith(".xls") else "openpyxl"
        xl  = pd.ExcelFile(io.BytesIO(raw), engine=eng)
        inserted, skipped = 0, 0
        for sheet in xl.sheet_names:
            df = xl.parse(sheet, header=0)
            df = norm_columnas(df)
            df = df[[c for c in df.columns if not c.startswith("Unnamed")]]
            if "DETALLE" not in df.columns: continue
            for _, row in df.iterrows():
                try:
                    raw_det = safe(row.get("DETALLE")) or safe(row.get("TIPO"))
                    if not raw_det: skipped += 1; continue
                    detalle, obs_extra = split_detalle_obs(raw_det)
                    obs_col   = safe(row.get("OBSERVACIONES O&M"))
                    obs_final = "\n\n".join(filter(None, [obs_col, obs_extra])) or None
                    resp_ing  = title_case(row.get("RESPONSABLE ING"))
                    if resp_ing:
                        await s.execute(text("INSERT INTO cfg_responsables (nombre) VALUES (:n) ON CONFLICT DO NOTHING"), {"n": resp_ing})
                    params = {
                        "tipo":              title_case(row.get("TIPO")) or "Vulnerabilidad",
                        "detalle":           detalle,
                        "subgerencia":       norm_subgerencia(row.get("SUBGERENCIA")),
                        "area":              title_case(row.get("AREA")),
                        "solicitante":       title_case(row.get("SOLICITANTE")),
                        "responsable_om":    title_case(row.get("RESPONSABLE")),
                        "responsable_ing":   resp_ing,
                        "prioridad":         norm_prioridad(row.get("PRIORIDAD")),
                        "fecha_declaracion": safe_date(row.get("FECHA DECLARACION")),
                        "fecha_compromiso":  safe_date(row.get("FECHA COMPROMISO")),
                        "fecha_solucion":    safe_date(row.get("FECHA SOLUCION")),
                        "estado_om":         norm_estado(row.get("ESTADO O&M")),
                        "estado_ing":        safe(row.get("ESTADO ING")),
                        "condicion_ing":     safe(row.get("CONDICION ING")),
                        "obs_om":            obs_final,
                        "obs_ing":           safe(row.get("OBSERVACIONES ING")),
                    }
                    sql = ("INSERT INTO vulnerabilidades "
                        "(tipo,detalle,subgerencia,area,solicitante,responsable_om,responsable_ing,"
                        "prioridad,fecha_declaracion,fecha_compromiso,fecha_solucion,"
                        "estado_om,estado_ing,condicion_ing,obs_om,obs_ing) VALUES "
                        "(:tipo,:detalle,:subgerencia,:area,:solicitante,:responsable_om,:responsable_ing,"
                        ":prioridad,:fecha_declaracion,:fecha_compromiso,:fecha_solucion,"
                        ":estado_om,:estado_ing,:condicion_ing,:obs_om,:obs_ing)")
                    await s.execute(text("SAVEPOINT sp_row"))
                    await s.execute(text(sql), params)
                    inserted += 1
                except Exception:
                    await s.execute(text("ROLLBACK TO SAVEPOINT sp_row"))
                    skipped += 1
        await s.commit()
        return {"inserted": inserted, "skipped": skipped}
    except Exception as e:
        raise HTTPException(400, f"Error: {e}")

@app.get("/api/dashboard/responsables")
async def dashboard_responsables(s: AsyncSession = Depends(db)):
    rows = await s.execute(text("""
        SELECT responsable_ing,
               COUNT(*) as total,
               COUNT(*) FILTER (WHERE estado_om='SIN ACCION') as pendiente,
               COUNT(*) FILTER (WHERE estado_om='EN CURSO') as en_curso,
               COUNT(*) FILTER (WHERE estado_om='FINALIZADO') as finalizado
        FROM vulnerabilidades
        WHERE responsable_ing IS NOT NULL
        GROUP BY responsable_ing
        ORDER BY total DESC
        LIMIT 10
    """))
    return [dict(r._mapping) for r in rows]

@app.get("/api/dashboard/responsables/areas")
async def dashboard_responsables_areas(s: AsyncSession = Depends(db)):
    rows = await s.execute(text("""
        SELECT responsable_ing, area,
               COUNT(*) as total,
               COUNT(*) FILTER (WHERE estado_om='SIN ACCION') as pendiente,
               COUNT(*) FILTER (WHERE estado_om='EN CURSO') as en_curso,
               COUNT(*) FILTER (WHERE estado_om='FINALIZADO') as finalizado
        FROM vulnerabilidades
        WHERE responsable_ing IS NOT NULL AND area IS NOT NULL
        GROUP BY responsable_ing, area
        ORDER BY responsable_ing, total DESC
    """))
    result = {}
    for r in rows:
        resp = r.responsable_ing
        if resp not in result:
            result[resp] = []
        result[resp].append({
            "area": r.area,
            "total": r.total,
            "pendiente": r.pendiente,
            "en_curso": r.en_curso,
            "finalizado": r.finalizado
        })
    return result

@app.get("/api/dashboard/capex")
async def dashboard_capex(s: AsyncSession = Depends(db)):
    rows = await s.execute(text("""
        SELECT 
            COUNT(*) FILTER (WHERE requiere_capex=TRUE) as requiere,
            COUNT(*) FILTER (WHERE requiere_capex=TRUE AND estado_capex='APROBADO') as aprobado,
            COUNT(*) FILTER (WHERE requiere_capex=TRUE AND estado_capex='RECHAZADO') as rechazado,
            COUNT(*) FILTER (WHERE requiere_capex=TRUE AND estado_capex='PENDIENTE') as pendiente,
            COUNT(*) FILTER (WHERE requiere_capex=TRUE AND estado_capex IS NULL) as sin_estado
        FROM vulnerabilidades
    """))
    r = rows.fetchone()
    return {
        "requiere": r.requiere,
        "aprobado": r.aprobado,
        "rechazado": r.rechazado,
        "pendiente": r.pendiente + r.sin_estado
    }

@app.get("/api/dashboard/capex")
async def dashboard_capex(s: AsyncSession = Depends(db)):
    rows = await s.execute(text("""
        SELECT 
            COUNT(*) FILTER (WHERE requiere_capex=TRUE) as requiere,
            COUNT(*) FILTER (WHERE requiere_capex=TRUE AND estado_capex='APROBADO') as aprobado,
            COUNT(*) FILTER (WHERE requiere_capex=TRUE AND estado_capex='RECHAZADO') as rechazado,
            COUNT(*) FILTER (WHERE requiere_capex=TRUE AND estado_capex='PENDIENTE') as pendiente,
            COUNT(*) FILTER (WHERE requiere_capex=TRUE AND estado_capex IS NULL) as sin_estado
        FROM vulnerabilidades
    """))
    r = rows.fetchone()
    return {
        "requiere": r.requiere,
        "aprobado": r.aprobado,
        "rechazado": r.rechazado,
        "pendiente": r.pendiente + r.sin_estado
    }

@app.patch("/api/vulnerabilidades/{vid}/capex")
async def actualizar_capex(vid: int, body: dict, s: AsyncSession = Depends(db)):
    requiere = body.get("requiere_capex", False)
    estado   = body.get("estado_capex", None)
    await s.execute(text(
        "UPDATE vulnerabilidades SET requiere_capex=:r, estado_capex=:e WHERE id=:id"
    ), {"r": requiere, "e": estado, "id": vid})
    await s.commit()
    return {"message": "OK"}

@app.get("/api/capex")
async def listar_capex(s: AsyncSession = Depends(db)):
    rows = await s.execute(text("""
        SELECT id, tipo, detalle, subgerencia, area, responsable_ing,
               prioridad, estado_om, requiere_capex, estado_capex
        FROM vulnerabilidades
        WHERE requiere_capex = TRUE
        ORDER BY 
            CASE estado_capex 
                WHEN 'PENDIENTE' THEN 1 
                WHEN 'APROBADO' THEN 2 
                WHEN 'RECHAZADO' THEN 3 
                ELSE 0 END,
            prioridad
    """))
    return [dict(r._mapping) for r in rows]
